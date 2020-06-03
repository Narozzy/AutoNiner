import ast
import string
import matplotlib.pyplot as plt
import numpy as np
import datetime
import pandas as pd
import fire
import utils.excel_functions as ef
from openpyxl import Workbook, load_workbook
from collections import OrderedDict
from decimal import *

jobTypeToDateTimeColsMap = {
    'DOOR': ['start_time', 'end_time', 'tmestamp'],
    'QUESTIONS': ['date']
}

jobTypeColumnsToSave = {
    'DOOR': ['sensor_id', 'tmestamp', 'in_count', 'out_count'],
    'QUESTIONS': ['id', 'date', 'ip_address', 'internal_notes', 'entered_by', 'desk_location', 'question', 'question_type']
}

master_template_cols = ['Location', 'Time', 'DayOfWeek', 'Month', 'Day', 'Month+Day', 'Term', 'Calendar', 'Semester', 'AY']

master_template_wb = load_workbook(filename='utils/master_lookup.xlsx')
# master_template_wb = load_workbook(filename='autoniner/utils/master_lookup.xlsx')
semesterMap, academicMap, termMap, locationMap = {}, {}, {}, {} # Maps in place of VLOOKUPs

semester_ws = master_template_wb['SemesterLookUp']
academic_ws = master_template_wb['AcadYrLookUp']
term_ws = master_template_wb['TermLookUp']
location_ws = master_template_wb['LocationLookUp']

for r_idx, rval in enumerate(semester_ws.iter_rows(min_row=2, max_row=semester_ws.max_row), start=2):
    semesterMap.update({semester_ws.cell(row=r_idx, column=1).value : semester_ws.cell(row=r_idx, column=2).value})

for r_idx, rval in enumerate(academic_ws.iter_rows(min_row=2, max_row=academic_ws.max_row), start=2):
    academicMap.update({academic_ws.cell(row=r_idx, column=1).value : {'year': academic_ws.cell(row=r_idx, column=2).value, 'order': academic_ws.cell(row=r_idx, column=3).value}})
        
for r_idx, rval in enumerate(term_ws.iter_rows(min_row=2, max_row=term_ws.max_row), start=2):
    termMap.update({term_ws.cell(row=r_idx, column=1).value : term_ws.cell(row=r_idx, column=2).value})

for r_idx, rval in enumerate(location_ws.iter_rows(min_row=2, max_row=location_ws.max_row), start=2):
    locationMap.update({location_ws.cell(row=r_idx, column=1).value : location_ws.cell(row=r_idx, column=2).value})


job_transform_lambdas = {
    'Location': lambda x: locationMap[x['sensor_id']],
}

def door_job_transform(df):
    location_col = []
    term_col = []
    sem_col = []
    ay_col = []
    for v in df['sensor_id']:
        location_col.append(locationMap[v])
    df['Location'] = location_col
    df['Time'] = pd.DatetimeIndex(df['tmestamp']).hour
    df['DayOfWeek'] = df['tmestamp'].dt.day_name()
    df['Month'] = df['tmestamp'].dt.strftime('%b')
    for row in df.iterrows():
        term_col.append(termMap[row[1]['Month'] + ' ' + str(row[1]['tmestamp'].day)])
    df['Term'] = term_col
    df['CalendarYr'] = df['tmestamp'].dt.year
    for row in df.iterrows():
        sem_col.append(row[1]['Term'] + ' ' + str(row[1]['CalendarYr'])) 
    df['Semester'] = sem_col
    for v in df['Semester']:
        ay_col.append(academicMap[v]['year'])
    df['AcademicYr'] = ay_col
    return df

def question_job_transform(df):
    term_col = []
    sem_col = []
    ay_col = []
    df['Time'] = pd.DatetimeIndex(df['date']).hour
    df['DayOfWeek'] = df['date'].dt.day_name()
    df['Month'] = df['date'].dt.strftime('%b')
    for row in df.iterrows():
        term_col.append(termMap[row[1]['Month'] + ' ' + str(row[1]['date'].day)])
    df['Term'] = term_col
    df['CalendarYr'] = df['date'].dt.year
    for row in df.iterrows():
        sem_col.append(row[1]['Term'] + ' ' + str(row[1]['CalendarYr'])) 
    df['Semester'] = sem_col
    for v in df['Semester']:
        ay_col.append(academicMap[v]['year'])
    df['AcademicYr'] = ay_col
    return df

def door_job_transform_viz(df):
    location_col = []
    term_col = []
    for v in df['sensor_id']:
        location_col.append(locationMap[v])
    df['Location'] = location_col
    north_total = df.loc[df['Location'] == 'North Entrance']['in_count'].sum() - df.loc[df['Location'] == 'North Entrance']['out_count'].sum()
    south_total = df.loc[df['Location'] == 'South Entrance']['in_count'].sum() - df.loc[df['Location'] == 'South Entrance']['out_count'].sum()
    coffee_total = df.loc[df['Location'] == 'CoffeeShop']['in_count'].sum() - df.loc[df['Location'] == 'CoffeeShop']['out_count'].sum()
    return north_total, south_total, coffee_total

def csv_transform(query_set, task_type):
    df_data = pd.DataFrame.from_records(query_set)
    for col in jobTypeToDateTimeColsMap[task_type]:
        df_data = ef.dateFormat(df_data, col, '%Y-%m-%d')
    df_data = df_data[jobTypeColumnsToSave[task_type]]
    if task_type == 'DOOR':
        df_data = door_job_transform(df_data)
    elif task_type == 'QUESTIONS':
        df_data = question_job_transform(df_data)
    return df_data

def construct_visualization(instances, task_type: str):
    if task_type == 'DOOR':
        return construct_door_viz(instances)
    elif task_type == 'QUESTIONS':
        return construct_question_viz(instances)

def construct_door_viz(instances):
    df = pd.DataFrame.from_records(instances)
    n, s, c = door_job_transform_viz(df)
    fig = plt.figure()
    ax = fig.add_axes([0,0,1,1])
    locations = ['North Entrance', 'South Entrance', 'Coffee Shop']
    nums = [n, s, c]
    ax.bar(locations, nums)
    return fig

def construct_question_viz(instances):
    df = pd.DataFrame.from_records(instances)
    return df

def constructDataVisualizationString(instances):

    dataString = "["
    condensedList = condenseHourInstances(instances)

    for i, hourData in enumerate(condensedList):
        if (i + 1) == len(condensedList):
            dataString += hourData.__str__()
        else:
            dataString += ( hourData.__str__() + ", " )

    dataString += "]"

    return dataString

def condenseHourInstances(instances):

    class monthlyHourData:
        hour = 0
        day = ""
        month = ""
        year = ""
        north_count = 0
        south_count = 0
        coffee_count = 0
        monthDay = ""

        def __init__(self, hour, day, month, year, n_count, s_count, c_count, monthDay ):
            self.hour = hour
            self.day = day
            self.month = month
            self.year = year
            self.north_count = n_count
            self.south_count = s_count
            self.coffee_count = c_count
            self.monthDay = monthDay

        def __str__(self):
            return '{ "hour":"' + str(self.hour) + '", ' + \
                   '"day":"' + self.day + '", ' + \
                   '"monthDay":"' + self.monthDay + '", ' + \
                   '"month":"' + self.month + '", ' + \
                   '"year":"' + self.year + '", ' + \
                   '"north_c":"' + str(self.north_count) + '", ' + \
                   '"south_c":"' + str(self.south_count) + '", ' + \
                   '"coffee_c":"' + str(self.coffee_count) + '"' + ' }'

    dictHourData = dict()

    for instance in instances:

        if "Entrance" or "Coffee" in instance["sensor_group"]:

            date = datetime.datetime.utcfromtimestamp((instance["tmestamp"] - 25569) * Decimal(86400.0))
            dateIntersection = date.strftime("%H-%a-%b-%Y")

            if dateIntersection not in dictHourData:

                north_count = 0
                south_count = 0
                coffee_count = 0

                # Assign value of intake to correct entrance
                if "North" in instance["sensor_group"]:
                    north_count = instance["in_count"]
                elif "SE" in instance["sensor_group"]:
                    south_count = instance["in_count"]
                elif "Coffee" in instance["sensor_group"]:
                    coffee_count = instance["in_count"]

                dictHourData[dateIntersection] = monthlyHourData( date.hour, date.strftime("%w"),
                                                                  str(date.strftime("%B")), str(date.strftime("%Y")),
                                                                  north_count, south_count, coffee_count, date.strftime("%d"))
            else: #update record to include new intake data

                tempHourData = dictHourData[dateIntersection]

                if "North" in instance["sensor_group"]:
                    tempHourData.north_count += instance["in_count"]
                elif "SE" in instance["sensor_group"]:
                    tempHourData.south_count += instance["in_count"]
                elif "Coffee" in instance["sensor_group"]:
                    tempHourData.coffee_count += instance["in_count"]

                dictHourData[dateIntersection] = tempHourData

    tempArray = list(dictHourData.values())

    return tempArray

